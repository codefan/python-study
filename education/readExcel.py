import openpyxl
from bs4 import BeautifulSoup
import urllib.request
import urllib.parse

wb = openpyxl.load_workbook('C:\\Users\\codef\\Documents\\temp\\words_f_z.xlsx')
sh = wb['Sheet1']
for i in range(751,1594):
    ce = sh.cell(row = i,column = 2)
    if ce and ce.value and len(ce.value) > 3:
        ce2 = sh.cell(row = i,column = 3)
        if not( ce2 and ce2.value and len(ce2.value) > 3):
            url ="http://dict.youdao.com/search?q=" + urllib.parse.quote(ce.value)
            print(url)
            request = urllib.request.Request(url)
            soup = BeautifulSoup(urllib.request.urlopen(url=request),features='html.parser')
            cells = soup.select(".phonetic")
            if cells and len(cells) > 0 :
                yinbiao = cells[0].text
                sh.cell(row = i,column = 3).value = yinbiao
            if cells and len(cells) > 1 :
                yinbiao = soup.select(".phonetic")[1].text
                sh.cell(row = i,column = 4).value = yinbiao
            # print( soup.title)
            cells = soup.select(".trans-container")
            if cells and len(cells) > 0 :
                yinbiao = cells[0].text
                if yinbiao and len(yinbiao) > 0:
                    sh.cell(row = i,column = 5).value = yinbiao
    if i % 10 == 0:
        print(i)
    if i % 50 == 0:
        wb.save('C:\\Users\\codef\\Documents\\temp\\words_f_z_'+str(i)+'.xlsx')
wb.save('C:\\Users\\codef\\Documents\\temp\\words_f_z2.xlsx')
wb.close()